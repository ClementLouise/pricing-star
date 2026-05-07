"""
Bulk XLSX import — parse, validate, plan, execute.

Round-trip with audit_pack.py: same sheet names, same column headers.
The 3 output-only sheets (Simulations, AuditLog, Methodology) are ignored
silently if present in the uploaded file.
"""
from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from typing import Any

import openpyxl
from pydantic import ValidationError

from app.engine.constants import COUNTRY_NAMES
from app.schemas.asset import AssetCreate, AssetUpdate
from app.schemas.scenario import (
    CountryDataInput,
    GenerousConfig,
    GlobeConfig,
    GuardConfig,
    LeversConfig,
    RegulationsConfig,
    ScenarioCreate,
)
from app.services.audit_pack import (
    COUNTRY_DATA_COLS,
    G2N_COLS,
    LEVER_COLS,
    OVERVIEW_KEYS,
    SCENARIO_COLS,
    VALID_LEVER_TYPES,
    _build_asset_full_xlsx,
    _header_row,
)

# Output-only sheets: present in audit pack export, ignored on import
IGNORED_SHEETS = frozenset({"Simulations", "AuditLog", "Methodology",
                             "Summary", "Cascade Prices", "Yearly Breakdown", "MFN Analysis"})


class ImportMode(str, Enum):
    DRY_RUN = "dry_run"
    CREATE_NEW = "create_new"
    UPDATE_EXISTING = "update_existing"


@dataclass
class ImportError:
    sheet: str
    row: int | None
    column: str | None
    code: str
    message: str
    raw_value: Any = None


@dataclass
class ParsedScenario:
    name: str
    description: str | None
    is_baseline: bool
    regulations: RegulationsConfig
    levers: LeversConfig
    country_data: list[tuple[str, CountryDataInput]]
    cascade_config: dict


@dataclass
class ParsedAsset:
    asset_payload: AssetCreate | AssetUpdate
    asset_id: uuid.UUID | None
    expected_updated_at: datetime | None
    scenarios: list[ParsedScenario]


@dataclass
class ImportPlan:
    mode: ImportMode
    parsed: ParsedAsset | None
    errors: list[ImportError]
    warnings: list[ImportError] = field(default_factory=list)
    summary: dict = field(default_factory=dict)


# ── Cell helpers ──────────────────────────────────────────────────────────────

def _to_bool(val: Any, default: bool = False) -> bool:
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    return s in ("true", "1", "yes", "oui")


def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _to_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _col_map(ws: Any) -> dict[str, int]:
    """Return {header_name: 1-based column index} from row 1."""
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value is not None:
                headers[str(cell.value).strip()] = cell.column
    return headers


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_overview_sheet(
    ws: Any,
    mode: ImportMode,
    target_asset_id: uuid.UUID | None,
    errors: list[ImportError],
) -> tuple[AssetCreate | AssetUpdate | None, uuid.UUID | None, datetime | None]:
    """Parse key/value Overview sheet. Returns (payload, asset_id, expected_updated_at)."""
    kv: dict[str, tuple[Any, int]] = {}
    for row in ws.iter_rows(min_row=1):
        key_cell = row[0] if row else None
        val_cell = row[1] if len(row) > 1 else None
        if key_cell and key_cell.value is not None:
            k = str(key_cell.value).strip()
            if k and k != "key":
                kv[k] = (val_cell.value if val_cell else None, key_cell.row)

    def _get(key: str) -> Any:
        return kv.get(key, (None, None))[0]

    def _row(key: str) -> int | None:
        return kv.get(key, (None, None))[1]

    # Asset ID and OCC token
    asset_id: uuid.UUID | None = None
    expected_updated_at: datetime | None = None

    raw_id = _get("asset_id")
    if raw_id:
        try:
            asset_id = uuid.UUID(str(raw_id).strip())
        except ValueError:
            errors.append(ImportError(
                sheet="Overview", row=_row("asset_id"), column="B",
                code="INVALID_UUID", message="asset_id n'est pas un UUID valide",
                raw_value=raw_id,
            ))

    if mode == ImportMode.UPDATE_EXISTING:
        if asset_id is None:
            errors.append(ImportError(
                sheet="Overview", row=None, column="B",
                code="MISSING_ASSET_ID",
                message="asset_id requis dans Overview pour le mode update_existing",
            ))
        elif target_asset_id and asset_id != target_asset_id:
            errors.append(ImportError(
                sheet="Overview", row=_row("asset_id"), column="B",
                code="ASSET_ID_MISMATCH",
                message=f"Le fichier cible l'asset {asset_id} mais l'endpoint cible {target_asset_id}",
                raw_value=str(raw_id),
            ))
        raw_ts = _get("updated_at")
        if raw_ts:
            try:
                expected_updated_at = datetime.fromisoformat(str(raw_ts).strip())
            except ValueError:
                pass  # OCC token optional — will skip OCC check if missing

    # Build asset fields dict
    asset_data: dict[str, Any] = {}
    _str_fields = {"name", "therapeutic_area", "modality", "indication"}
    _float_fields = {"us_list_price", "us_net_share", "cogs_percent", "discount_rate",
                     "peak_capture_rate", "part_b_share"}
    _int_fields = {"launch_year", "peak_year", "loe_year",
                   "us_patient_population", "ex_us_patient_population", "ramp_years"}

    for f in _str_fields:
        v = _to_str(_get(f))
        if v is not None:
            asset_data[f] = v

    for f in _float_fields:
        v = _to_float(_get(f))
        if v is not None:
            asset_data[f] = v

    for f in _int_fields:
        v = _to_int(_get(f))
        if v is not None:
            asset_data[f] = v

    if mode == ImportMode.CREATE_NEW:
        if not asset_data.get("name"):
            errors.append(ImportError(
                sheet="Overview", row=_row("name"), column="B",
                code="MISSING_REQUIRED_FIELD",
                message="Le champ 'name' est requis dans Overview",
            ))
            return None, asset_id, expected_updated_at
        try:
            payload: AssetCreate | AssetUpdate = AssetCreate(**asset_data)
        except ValidationError as exc:
            for e in exc.errors():
                loc = e["loc"][0] if e["loc"] else "?"
                errors.append(ImportError(
                    sheet="Overview", row=_row(str(loc)), column="B",
                    code=e["type"].upper().replace("-", "_"),
                    message=f"Champ '{loc}' : {e['msg']}",
                    raw_value=asset_data.get(str(loc)),
                ))
            return None, asset_id, expected_updated_at
    else:
        payload = AssetUpdate(**asset_data)

    return payload, asset_id, expected_updated_at


def _parse_scenarios_sheet(ws: Any, errors: list[ImportError]) -> list[dict[str, Any]]:
    cols = _col_map(ws)
    if "scenario_name" not in cols:
        errors.append(ImportError(
            sheet="Scenarios", row=1, column="scenario_name",
            code="MISSING_COLUMN", message="Colonne 'scenario_name' manquante dans Scenarios",
        ))
        return []

    results: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name_val = row[cols["scenario_name"] - 1] if cols["scenario_name"] - 1 < len(row) else None
        if name_val is None:
            continue
        name = _to_str(name_val)
        if not name:
            continue

        def _cell(col: str) -> Any:
            idx = cols.get(col)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        results.append({
            "name": name,
            "description": _to_str(_cell("description")),
            "is_baseline": _to_bool(_cell("is_baseline"), False),
            "regulations": {
                "generous": {
                    "active": _to_bool(_cell("generous_active"), False),
                    "year": _to_int(_cell("generous_year")),
                    "medicaid_share": _to_float(_cell("generous_medicaid_share")) or 0.07,
                },
                "guard": {
                    "active": _to_bool(_cell("guard_active"), False),
                    "year": _to_int(_cell("guard_year")),
                    "submit_method_ii": _to_bool(_cell("guard_submit_method_ii"), False),
                    "phase_in": _to_float(_cell("guard_phase_in")),
                },
                "globe": {
                    "active": _to_bool(_cell("globe_active"), False),
                    "year": _to_int(_cell("globe_year")),
                    "submit_method_ii": _to_bool(_cell("globe_submit_method_ii"), False),
                    "phase_in": _to_float(_cell("globe_phase_in")),
                },
            },
            "cascade_config": {
                "enabled": _to_bool(_cell("cascade_enabled"), True),
                "max_iter": _to_int(_cell("cascade_max_iter")) or 10,
            },
        })
    return results


def _parse_country_data_sheet(
    ws: Any,
    valid_scenario_names: set[str],
    errors: list[ImportError],
) -> dict[str, list[tuple[str, dict]]]:
    cols = _col_map(ws)
    for req in ("scenario_name", "country_code"):
        if req not in cols:
            errors.append(ImportError(
                sheet="CountryData", row=1, column=req,
                code="MISSING_COLUMN", message=f"Colonne '{req}' manquante dans CountryData",
            ))
            return {}

    results: dict[str, list[tuple[str, dict]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(col: str) -> Any:
            idx = cols.get(col)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        sc_name = _to_str(_cell("scenario_name"))
        cc = _to_str(_cell("country_code"))
        if sc_name is None or cc is None:
            continue

        row_num = None  # can't get row number in values_only mode easily
        if sc_name not in valid_scenario_names:
            errors.append(ImportError(
                sheet="CountryData", row=row_num, column="scenario_name",
                code="UNKNOWN_SCENARIO_REF",
                message=f"Scénario '{sc_name}' introuvable dans l'onglet Scenarios",
                raw_value=sc_name,
            ))
            continue
        if cc not in COUNTRY_NAMES:
            errors.append(ImportError(
                sheet="CountryData", row=row_num, column="country_code",
                code="INVALID_COUNTRY_CODE",
                message=f"Code pays '{cc}' invalide. Codes acceptés : {sorted(COUNTRY_NAMES)}",
                raw_value=cc,
            ))
            continue

        cd_dict = {
            "list_price": _to_float(_cell("list_price")),
            "net_price": _to_float(_cell("net_price")),
            "volume": _to_float(_cell("volume")),
            "launched": _to_bool(_cell("launched"), False),
            "launch_year": _to_int(_cell("launch_year")),
            "withdrawn": _to_bool(_cell("withdrawn"), False),
            "g2n_ratio": _to_float(_cell("g2n_ratio")),
        }
        results.setdefault(sc_name, []).append((cc, cd_dict))
    return results


def _parse_g2n_time_series_sheet(
    ws: Any,
    valid_scenario_names: set[str],
    errors: list[ImportError],
) -> dict[str, dict[str, dict[int, float]]]:
    cols = _col_map(ws)
    for req in ("scenario_name", "country_code", "year", "g2n_value"):
        if req not in cols:
            return {}  # sheet is optional; missing columns → silently skip

    results: dict[str, dict[str, dict[int, float]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(col: str) -> Any:
            idx = cols.get(col)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        sc_name = _to_str(_cell("scenario_name"))
        cc = _to_str(_cell("country_code"))
        year_val = _to_int(_cell("year"))
        g2n_val = _to_float(_cell("g2n_value"))

        if None in (sc_name, cc, year_val, g2n_val):
            continue
        if sc_name not in valid_scenario_names:
            continue  # already caught in CountryData
        if cc not in COUNTRY_NAMES:
            continue

        if not (2020 <= year_val <= 2075):  # type: ignore[operator]
            errors.append(ImportError(
                sheet="G2N_TimeSeries", row=None, column="year",
                code="G2N_YEAR_OUT_OF_RANGE",
                message=f"Année G2N {year_val} hors de la plage [2020, 2075]",
                raw_value=year_val,
            ))
            continue
        if not (0 < g2n_val <= 1):  # type: ignore[operator]
            errors.append(ImportError(
                sheet="G2N_TimeSeries", row=None, column="g2n_value",
                code="INVALID_G2N",
                message=f"Valeur G2N {g2n_val} doit être dans (0, 1]",
                raw_value=g2n_val,
            ))
            continue

        results.setdefault(sc_name, {}).setdefault(cc, {})[year_val] = g2n_val  # type: ignore[index]
    return results


def _parse_levers_sheet(
    ws: Any,
    valid_scenario_names: set[str],
    errors: list[ImportError],
) -> dict[str, LeversConfig]:
    cols = _col_map(ws)
    if "scenario_name" not in cols or "lever_type" not in cols:
        return {}  # optional sheet

    raw: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(col: str) -> Any:
            idx = cols.get(col)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        sc_name = _to_str(_cell("scenario_name"))
        lever_type = _to_str(_cell("lever_type"))
        if sc_name is None or lever_type is None:
            continue
        if sc_name not in valid_scenario_names:
            continue
        if lever_type not in VALID_LEVER_TYPES:
            errors.append(ImportError(
                sheet="Levers", row=None, column="lever_type",
                code="INVALID_LEVER_TYPE",
                message=f"Type de levier '{lever_type}' inconnu. Valeurs : {sorted(VALID_LEVER_TYPES)}",
                raw_value=lever_type,
            ))
            continue

        d = raw.setdefault(sc_name, {
            "withdrawals": [], "price_floors": {}, "delayed_launches": {},
            "de_opt_in": False, "gr_clawback_stress": False,
        })
        country = _to_str(_cell("country_code")) or ""
        value = _cell("value")

        if lever_type == "withdrawal" and country:
            if _to_bool(value, True) and country not in d["withdrawals"]:
                d["withdrawals"].append(country)
        elif lever_type == "price_floor" and country:
            fv = _to_float(value)
            if fv is not None:
                d["price_floors"][country] = fv
        elif lever_type == "delayed_launch" and country:
            iv = _to_int(value)
            if iv is not None:
                d["delayed_launches"][country] = iv
        elif lever_type == "de_opt_in":
            d["de_opt_in"] = _to_bool(value, False)
        elif lever_type == "gr_clawback_stress":
            d["gr_clawback_stress"] = _to_bool(value, False)

    return {name: LeversConfig(**d) for name, d in raw.items()}


# ── Cross-reference + business validation ────────────────────────────────────

def _validate_cross_references(
    scenarios: list[dict[str, Any]],
    country_data_by_sc: dict[str, list[tuple[str, dict]]],
    g2n_by_sc: dict[str, dict[str, dict[int, float]]],
    errors: list[ImportError],
    warnings: list[ImportError],
) -> None:
    baseline_count = sum(1 for s in scenarios if s.get("is_baseline"))
    if baseline_count > 1:
        errors.append(ImportError(
            sheet="Scenarios", row=None, column="is_baseline",
            code="MULTIPLE_BASELINES",
            message=f"{baseline_count} scénarios marqués is_baseline=TRUE. Un seul autorisé.",
        ))
    if not scenarios:
        warnings.append(ImportError(
            sheet="Scenarios", row=None, column=None,
            code="NO_SCENARIOS",
            message="Aucun scénario dans le fichier. L'import créera l'asset sans scénario.",
        ))

    # G2N references must point to a country in CountryData
    for sc_name, country_g2n in g2n_by_sc.items():
        cd_countries = {cc for cc, _ in country_data_by_sc.get(sc_name, [])}
        for cc in country_g2n:
            if cc not in cd_countries:
                errors.append(ImportError(
                    sheet="G2N_TimeSeries", row=None, column="country_code",
                    code="UNKNOWN_COUNTRY_REF",
                    message=f"Pays '{cc}' dans G2N_TimeSeries absent de CountryData pour le scénario '{sc_name}'",
                    raw_value=cc,
                ))


def _detect_real_pricing_data(
    scenarios: list[dict[str, Any]],
    country_data_by_sc: dict[str, list[tuple[str, dict]]],
) -> list[ImportError]:
    """Heuristic: if any country list_price > $100K → trial-mode warning."""
    for cd_list in country_data_by_sc.values():
        for _, cd in cd_list:
            lp = cd.get("list_price")
            if lp and lp > 100_000:
                return [ImportError(
                    sheet="CountryData", row=None, column="list_price",
                    code="REAL_PRICING_DETECTED",
                    message=(
                        "Des prix >100 000$ ont été détectés. En mode Trial, "
                        "seules des données illustratives sont autorisées."
                    ),
                    raw_value=lp,
                )]
    return []


# ── Summary ───────────────────────────────────────────────────────────────────

def _summarize(scenarios: list[dict], cd_by_sc: dict) -> dict[str, Any]:
    total_cd = sum(len(v) for v in cd_by_sc.values())
    return {
        "scenario_count": len(scenarios),
        "country_data_count": total_cd,
        "scenario_names": [s.get("name") for s in scenarios],
    }


# ── Main entry point ──────────────────────────────────────────────────────────

def parse_xlsx(
    file_bytes: bytes,
    mode: ImportMode,
    target_asset_id: uuid.UUID | None = None,
) -> ImportPlan:
    """Parse XLSX bytes into a validated ImportPlan. NO database calls."""
    errors: list[ImportError] = []
    warnings: list[ImportError] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return ImportPlan(
            mode=mode, parsed=None,
            errors=[ImportError(
                sheet="(file)", row=None, column=None,
                code="INVALID_XLSX",
                message=f"Fichier XLSX invalide ou corrompu : {exc}",
            )],
        )

    sheet_names = set(wb.sheetnames)
    if mode == ImportMode.CREATE_NEW and "Overview" not in sheet_names:
        return ImportPlan(
            mode=mode, parsed=None,
            errors=[ImportError(
                sheet="(file)", row=None, column=None,
                code="MISSING_REQUIRED_SHEET",
                message="L'onglet 'Overview' est requis pour le mode create_new",
            )],
        )

    # Parse each sheet
    asset_payload: AssetCreate | AssetUpdate | None = None
    asset_id: uuid.UUID | None = None
    expected_updated_at: datetime | None = None

    if "Overview" in sheet_names:
        asset_payload, asset_id, expected_updated_at = _parse_overview_sheet(
            wb["Overview"], mode, target_asset_id, errors
        )

    scenarios_data: list[dict[str, Any]] = []
    if "Scenarios" in sheet_names:
        scenarios_data = _parse_scenarios_sheet(wb["Scenarios"], errors)

    valid_names = {s["name"] for s in scenarios_data if s.get("name")}

    cd_by_sc: dict[str, list[tuple[str, dict]]] = {}
    if "CountryData" in sheet_names:
        cd_by_sc = _parse_country_data_sheet(wb["CountryData"], valid_names, errors)

    g2n_by_sc: dict[str, dict[str, dict[int, float]]] = {}
    if "G2N_TimeSeries" in sheet_names:
        g2n_by_sc = _parse_g2n_time_series_sheet(wb["G2N_TimeSeries"], valid_names, errors)

    levers_by_sc: dict[str, LeversConfig] = {}
    if "Levers" in sheet_names:
        levers_by_sc = _parse_levers_sheet(wb["Levers"], valid_names, errors)

    if errors:
        return ImportPlan(mode=mode, parsed=None, errors=errors, warnings=warnings,
                          summary=_summarize(scenarios_data, cd_by_sc))

    # Merge G2N time series into country data dicts
    for sc_name, cd_list in cd_by_sc.items():
        g2n_for_sc = g2n_by_sc.get(sc_name, {})
        for cc, cd_dict in cd_list:
            ts = g2n_for_sc.get(cc)
            if ts:
                cd_dict["g2n_time_series"] = {str(y): v for y, v in ts.items()}

    # Cross-reference validation
    _validate_cross_references(scenarios_data, cd_by_sc, g2n_by_sc, errors, warnings)
    if errors:
        return ImportPlan(mode=mode, parsed=None, errors=errors, warnings=warnings,
                          summary=_summarize(scenarios_data, cd_by_sc))

    # Build ParsedScenario objects with Pydantic validation
    parsed_scenarios: list[ParsedScenario] = []
    for sc_data in scenarios_data:
        name = sc_data["name"]
        try:
            regs = RegulationsConfig(
                generous=GenerousConfig(**sc_data["regulations"]["generous"]),
                guard=GuardConfig(**sc_data["regulations"]["guard"]),
                globe=GlobeConfig(**sc_data["regulations"]["globe"]),
            )
        except ValidationError as exc:
            for e in exc.errors():
                errors.append(ImportError(
                    sheet="Scenarios", row=None, column=str(e["loc"][-1] if e["loc"] else "?"),
                    code=e["type"].upper().replace("-", "_"),
                    message=f"Scénario '{name}' regulations : {e['msg']}",
                ))
            continue

        country_data: list[tuple[str, CountryDataInput]] = []
        for cc, cd_dict in cd_by_sc.get(name, []):
            try:
                country_data.append((cc, CountryDataInput(**cd_dict)))
            except ValidationError as exc:
                for e in exc.errors():
                    errors.append(ImportError(
                        sheet="CountryData", row=None,
                        column=str(e["loc"][-1] if e["loc"] else "?"),
                        code=e["type"].upper().replace("-", "_"),
                        message=f"Pays '{cc}', scénario '{name}' : {e['msg']}",
                        raw_value=cd_dict.get(str(e["loc"][-1]) if e["loc"] else ""),
                    ))

        levers = levers_by_sc.get(name, LeversConfig())
        parsed_scenarios.append(ParsedScenario(
            name=name,
            description=sc_data.get("description"),
            is_baseline=sc_data.get("is_baseline", False),
            regulations=regs,
            levers=levers,
            country_data=country_data,
            cascade_config=sc_data.get("cascade_config", {}),
        ))

    if errors:
        return ImportPlan(mode=mode, parsed=None, errors=errors, warnings=warnings,
                          summary=_summarize(scenarios_data, cd_by_sc))

    assert asset_payload is not None
    parsed = ParsedAsset(
        asset_payload=asset_payload,
        asset_id=asset_id,
        expected_updated_at=expected_updated_at,
        scenarios=parsed_scenarios,
    )

    warnings.extend(_detect_real_pricing_data(scenarios_data, cd_by_sc))
    return ImportPlan(
        mode=mode, parsed=parsed, errors=[], warnings=warnings,
        summary=_summarize(scenarios_data, cd_by_sc),
    )


# ── DB execution ──────────────────────────────────────────────────────────────

async def execute_import_plan(
    plan: ImportPlan,
    db: object,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
    tier: str = "production",
    trial_expires_at: datetime | None = None,
) -> dict[str, Any]:
    """
    Execute a validated ImportPlan against the database. All-or-nothing.
    Caller handles audit logging and commit.
    """
    from sqlalchemy.ext.asyncio import AsyncSession

    from app.api._occ import check_occ
    from app.repos.asset import AssetRepo
    from app.repos.scenario import CountryDataRepo, ScenarioRepo
    from app.services.trial_limits import assert_can_create_asset, assert_can_create_scenario

    assert isinstance(db, AsyncSession)
    if plan.errors or plan.parsed is None:
        raise ValueError("Cannot execute plan with validation errors")

    parsed = plan.parsed
    asset_repo = AssetRepo(db)
    scenario_repo = ScenarioRepo(db)
    cd_repo = CountryDataRepo(db)

    scenario_ids: list[uuid.UUID] = []
    cd_count = 0

    if plan.mode == ImportMode.CREATE_NEW:
        assert isinstance(parsed.asset_payload, AssetCreate)
        await assert_can_create_asset(tenant_id, tier, trial_expires_at, db)
        asset = await asset_repo.create(tenant_id, user_id, parsed.asset_payload)

        for ps in parsed.scenarios:
            await assert_can_create_scenario(asset.id, tenant_id, tier, trial_expires_at, db)
            sc_payload = ScenarioCreate(
                name=ps.name,
                description=ps.description,
                is_baseline=ps.is_baseline,
                regulations=ps.regulations,
                levers=ps.levers,
                cascade_config=ps.cascade_config,
            )
            scenario = await scenario_repo.create(asset.id, tenant_id, user_id, sc_payload)
            scenario_ids.append(scenario.id)
            for cc, cd_input in ps.country_data:
                await cd_repo.upsert(scenario.id, tenant_id, cc, cd_input)
                cd_count += 1

        return {
            "asset_id": str(asset.id),
            "scenario_ids": [str(s) for s in scenario_ids],
            "country_data_count": cd_count,
        }

    # UPDATE_EXISTING
    assert isinstance(parsed.asset_payload, AssetUpdate)
    asset = await asset_repo.get(parsed.asset_id, tenant_id)
    if asset is None:
        raise ValueError(f"Asset {parsed.asset_id} not found for tenant {tenant_id}")

    # OCC at asset level only (per decision: no per-scenario OCC in v1)
    parsed.asset_payload.expected_updated_at = parsed.expected_updated_at
    check_occ(asset.updated_at, parsed.expected_updated_at, parsed.asset_payload.force_override)
    asset = await asset_repo.update(asset, parsed.asset_payload)

    existing_scenarios = await scenario_repo.list(asset.id, tenant_id)
    existing_by_name = {s.name.lower(): s for s in existing_scenarios}

    for ps in parsed.scenarios:
        existing = existing_by_name.get(ps.name.lower())
        if existing:
            from app.schemas.scenario import ScenarioUpdate
            sc_update = ScenarioUpdate(
                name=ps.name,
                description=ps.description,
                is_baseline=ps.is_baseline,
                regulations=ps.regulations,
                levers=ps.levers,
                cascade_config=ps.cascade_config,
            )
            scenario = await scenario_repo.update(existing, sc_update)
        else:
            await assert_can_create_scenario(asset.id, tenant_id, tier, trial_expires_at, db)
            sc_payload = ScenarioCreate(
                name=ps.name, description=ps.description, is_baseline=ps.is_baseline,
                regulations=ps.regulations, levers=ps.levers, cascade_config=ps.cascade_config,
            )
            scenario = await scenario_repo.create(asset.id, tenant_id, user_id, sc_payload)
        scenario_ids.append(scenario.id)
        for cc, cd_input in ps.country_data:
            await cd_repo.upsert(scenario.id, tenant_id, cc, cd_input)
            cd_count += 1

    return {
        "asset_id": str(asset.id),
        "scenario_ids": [str(s) for s in scenario_ids],
        "country_data_count": cd_count,
    }


# ── Template generation ───────────────────────────────────────────────────────

def build_import_template(asset_name: str = "My Asset") -> bytes:
    """Generate an empty XLSX import template with all 5 sheets and a sample row."""
    from openpyxl.styles import Font, PatternFill

    _HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    _HDR_FILL = PatternFill("solid", fgColor="1A1A2E")

    def _header(ws: Any, cols: list[str]) -> None:
        for i, label in enumerate(cols, start=1):
            c = ws.cell(row=1, column=i, value=label)
            c.font = _HDR_FONT
            c.fill = _HDR_FILL

    sample_meta: dict[str, Any] = {
        "asset_id": "", "updated_at": "",
        "name": asset_name, "therapeutic_area": "", "modality": "", "indication": "",
        "us_list_price": 200000, "us_net_share": 0.75,
        "launch_year": 2027, "peak_year": 2031, "loe_year": 2039,
        "cogs_percent": 0.15, "discount_rate": 0.10,
        "us_patient_population": 50000, "ex_us_patient_population": 30000,
        "peak_capture_rate": 0.50, "part_b_share": 0.50, "ramp_years": 4,
    }
    sample_sc = {
        "name": "Scenario 1", "description": "", "is_baseline": True,
        "regulations": {
            "generous": {"active": False, "year": None, "medicaid_share": 0.07},
            "guard": {"active": False, "year": None, "submit_method_ii": False, "phase_in": None},
            "globe": {"active": False, "year": None, "submit_method_ii": False, "phase_in": None},
        },
        "cascade_config": {"enabled": True, "max_iter": 10},
        "levers": {
            "withdrawals": [], "price_floors": {}, "delayed_launches": {},
            "de_opt_in": False, "gr_clawback_stress": False,
        },
        "country_data": [{
            "country_code": "US", "list_price": 200000, "net_price": 150000,
            "volume": 0.5, "launched": True, "launch_year": 2027,
            "withdrawn": False, "g2n_ratio": 0.75, "g2n_time_series": {"2028": 0.73, "2029": 0.71},
        }],
    }

    return _build_asset_full_xlsx(sample_meta, [sample_sc])
