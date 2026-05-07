"""
Audit Pack + GDPR Data Export service.

Builds in-memory ZIP archives containing SOX-grade audit JSON, pricing XLSX,
and plain-text README for simulations, assets, and full tenant GDPR exports.
"""
from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_EXPORT_BYTES = 300 * 1024 * 1024  # 300 MB hard limit


class ExportTooLargeError(Exception):
    """Raised when a generated export exceeds MAX_EXPORT_BYTES."""


# ── Round-trip import/export column definitions ───────────────────────────────
# Canonical schema shared with asset_import.py. Column order here determines
# column order in both the export (asset_full.xlsx) and the import template.

OVERVIEW_KEYS: list[str] = [
    "asset_id", "updated_at", "name", "therapeutic_area", "modality", "indication",
    "us_list_price", "us_net_share", "launch_year", "peak_year", "loe_year",
    "cogs_percent", "discount_rate", "us_patient_population", "ex_us_patient_population",
    "peak_capture_rate", "part_b_share", "ramp_years",
]

SCENARIO_COLS: list[str] = [
    "scenario_name", "description", "is_baseline",
    "generous_active", "generous_year", "generous_medicaid_share",
    "guard_active", "guard_year", "guard_submit_method_ii", "guard_phase_in",
    "globe_active", "globe_year", "globe_submit_method_ii", "globe_phase_in",
    "cascade_enabled", "cascade_max_iter",
]

COUNTRY_DATA_COLS: list[str] = [
    "scenario_name", "country_code",
    "list_price", "net_price", "volume",
    "launched", "launch_year", "withdrawn", "g2n_ratio",
]

G2N_COLS: list[str] = ["scenario_name", "country_code", "year", "g2n_value"]

LEVER_COLS: list[str] = ["scenario_name", "lever_type", "country_code", "value"]

VALID_LEVER_TYPES: frozenset[str] = frozenset({
    "withdrawal", "price_floor", "delayed_launch", "de_opt_in", "gr_clawback_stress",
})


# ── XLSX builder ─────────────────────────────────────────────────────────────

_GOLD = "C9A84C"
_DARK = "1A1A2E"
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor=_DARK)
_TITLE_FONT = Font(bold=True, color=_GOLD, size=11)


def _header_row(ws: Any, cols: list[str], row: int = 1) -> None:
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18


def _autofit(ws: Any) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 40)


def build_simulation_xlsx(
    asset: dict[str, Any],
    scenario_name: str,
    sim: dict[str, Any],
) -> bytes:
    """
    Build a multi-sheet Excel workbook for a single simulation result.

    Sheets: Summary | Cascade Prices | Yearly Breakdown | Method I/II Detail
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    meta: list[tuple[str, Any]] = [
        ("Asset", asset.get("name")),
        ("Therapeutic Area", asset.get("therapeutic_area")),
        ("Modality", asset.get("modality")),
        ("Launch Year", asset.get("launch_year")),
        ("LOE Year", asset.get("patent_expiry") or asset.get("loe_year")),
        ("Discount Rate", f"{float(asset.get('discount_rate') or 0.10):.1%}"),
        ("Scenario", scenario_name),
        ("Engine Version", sim.get("engine_version")),
        ("Computed At", sim.get("computed_at")),
        ("", None),
        ("NPV (14-Year)", sim.get("npv")),
        ("Peak Revenue", sim.get("peak_revenue")),
        ("Method I Value", sim.get("method_i_value")),
        ("Method I Anchor", sim.get("method_i_anchor")),
        ("Method II Value", sim.get("method_ii_value")),
        ("Applicable Benchmark", sim.get("applicable_benchmark")),
        ("Per-Unit Rebate", sim.get("per_unit_rebate")),
        ("Effective US Net", sim.get("effective_us_net")),
        ("Cascade Converged", sim.get("cascade_converged")),
        ("Cascade Iterations", sim.get("cascade_iterations")),
    ]

    ws["A1"].value = "PRICING STAR — Audit Pack"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 20

    for row_idx, (label, value) in enumerate(meta, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row_idx, column=2, value=value)

    _autofit(ws)

    # ── Sheet 2: Cascade Prices ───────────────────────────────────────────
    ws2 = wb.create_sheet("Cascade Prices")
    _header_row(ws2, ["Country Code", "Price (USD)"])
    final_prices: dict[str, float] = sim.get("final_prices") or {}
    for row_idx, (code, price) in enumerate(sorted(final_prices.items()), start=2):
        ws2.cell(row=row_idx, column=1, value=code)
        ws2.cell(row=row_idx, column=2, value=price)
    _autofit(ws2)

    # ── Sheet 3: Yearly Breakdown ─────────────────────────────────────────
    ws3 = wb.create_sheet("Yearly Breakdown")
    _header_row(ws3, ["Year", "US Revenue", "Ex-US Revenue", "Total Net", "Effective US Net"])
    yearly: list[dict[str, Any]] = sim.get("yearly_breakdown") or []
    for row_idx, row in enumerate(yearly, start=2):
        ws3.cell(row=row_idx, column=1, value=row.get("year"))
        ws3.cell(row=row_idx, column=2, value=row.get("us_revenue"))
        ws3.cell(row=row_idx, column=3, value=row.get("ex_us_revenue"))
        ws3.cell(row=row_idx, column=4, value=row.get("total_net", (row.get("us_revenue") or 0) + (row.get("ex_us_revenue") or 0)))
        ws3.cell(row=row_idx, column=5, value=row.get("effective_us_net"))
    _autofit(ws3)

    # ── Sheet 4: Method I/II Detail ───────────────────────────────────────
    ws4 = wb.create_sheet("MFN Analysis")
    _header_row(ws4, ["Parameter", "Value"])
    mfn_rows = [
        ("Method I Value (USD)", sim.get("method_i_value")),
        ("Method I Anchor Country", sim.get("method_i_anchor")),
        ("Method II Value (USD)", sim.get("method_ii_value")),
        ("Applicable Benchmark (USD)", sim.get("applicable_benchmark")),
        ("Per-Unit Rebate (USD)", sim.get("per_unit_rebate")),
        ("Effective US Net Price (USD)", sim.get("effective_us_net")),
    ]
    for row_idx, (label, value) in enumerate(mfn_rows, start=2):
        ws4.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=10)
        ws4.cell(row=row_idx, column=2, value=value)
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── README builder ────────────────────────────────────────────────────────────

def build_readme_txt(
    asset_name: str,
    scenario_name: str,
    sim_id: str,
    generated_at: str,
    generated_by: str,
) -> str:
    return (
        f"PRICING STAR — Audit Pack\n"
        f"{'=' * 40}\n\n"
        f"Asset:       {asset_name}\n"
        f"Scenario:    {scenario_name}\n"
        f"Sim ID:      {sim_id}\n"
        f"Generated:   {generated_at}\n"
        f"Generated by (user ID): {generated_by}\n\n"
        f"Contents\n"
        f"--------\n"
        f"  audit.json   — SOX-grade audit document (CMS methodology + all intermediates)\n"
        f"  prices.xlsx  — Excel workbook: cascade prices, yearly breakdown, MFN analysis\n"
        f"  README.txt   — this file\n\n"
        f"Classification: Internal — Confidential\n"
        f"Do not share outside the pricing team without explicit approval.\n"
    )


# ── ZIP builders ──────────────────────────────────────────────────────────────

def build_simulation_pack(
    asset: dict[str, Any],
    scenario_name: str,
    sim: dict[str, Any],
    audit_json: dict[str, Any],
    generated_by: str,
) -> bytes:
    """
    ZIP containing audit.json + prices.xlsx + README.txt for one simulation.
    Raises ExportTooLargeError if the pack exceeds MAX_EXPORT_BYTES.
    """
    sim_id = sim.get("simulation_id") or sim.get("id") or "unknown"
    generated_at = datetime.now(timezone.utc).isoformat()

    xlsx_bytes = build_simulation_xlsx(asset, scenario_name, sim)
    readme_txt = build_readme_txt(
        asset_name=asset.get("name", "unknown"),
        scenario_name=scenario_name,
        sim_id=str(sim_id),
        generated_at=generated_at,
        generated_by=generated_by,
    )

    import json

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", readme_txt)
        zf.writestr("audit.json", json.dumps(audit_json, indent=2))
        zf.writestr("prices.xlsx", xlsx_bytes)

    raw = buf.getvalue()
    if len(raw) > MAX_EXPORT_BYTES:
        raise ExportTooLargeError(f"Pack size {len(raw):,} bytes exceeds limit {MAX_EXPORT_BYTES:,}")
    return raw


def _build_asset_full_xlsx(
    asset_meta: dict[str, Any],
    scenario_configs: list[dict[str, Any]],
) -> bytes:
    """
    Build the 5-sheet importable XLSX (asset_full.xlsx).
    Round-trip compatible with asset_import.parse_xlsx().
    asset_meta keys must include all OVERVIEW_KEYS fields.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview (key / value) ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    for col, label in enumerate(["key", "value"], start=1):
        c = ws1.cell(row=1, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL

    for row_idx, key in enumerate(OVERVIEW_KEYS, start=2):
        ws1.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=10)
        ws1.cell(row=row_idx, column=2, value=asset_meta.get(key))
    _autofit(ws1)

    # ── Sheet 2: Scenarios ────────────────────────────────────────────
    ws2 = wb.create_sheet("Scenarios")
    _header_row(ws2, SCENARIO_COLS)
    for row_idx, sc in enumerate(scenario_configs, start=2):
        regs = sc.get("regulations") or {}
        generous = regs.get("generous") or {}
        guard = regs.get("guard") or {}
        globe = regs.get("globe") or {}
        cascade = sc.get("cascade_config") or {}
        vals = [
            sc.get("name"), sc.get("description"), sc.get("is_baseline", False),
            generous.get("active", False), generous.get("year"), generous.get("medicaid_share", 0.07),
            guard.get("active", False), guard.get("year"), guard.get("submit_method_ii", False), guard.get("phase_in"),
            globe.get("active", False), globe.get("year"), globe.get("submit_method_ii", False), globe.get("phase_in"),
            cascade.get("enabled", True), cascade.get("max_iter", 10),
        ]
        for col_idx, v in enumerate(vals, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=v)
    _autofit(ws2)

    # ── Sheet 3: CountryData ──────────────────────────────────────────
    ws3 = wb.create_sheet("CountryData")
    _header_row(ws3, COUNTRY_DATA_COLS)
    row_idx = 2
    for sc in scenario_configs:
        for cd in sc.get("country_data") or []:
            vals = [
                sc.get("name"), cd.get("country_code"),
                cd.get("list_price"), cd.get("net_price"), cd.get("volume"),
                cd.get("launched", False), cd.get("launch_year"),
                cd.get("withdrawn", False), cd.get("g2n_ratio"),
            ]
            for col_idx, v in enumerate(vals, start=1):
                ws3.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
    _autofit(ws3)

    # ── Sheet 4: G2N_TimeSeries ───────────────────────────────────────
    ws4 = wb.create_sheet("G2N_TimeSeries")
    _header_row(ws4, G2N_COLS)
    row_idx = 2
    for sc in scenario_configs:
        for cd in sc.get("country_data") or []:
            for year_str, g2n_val in sorted((cd.get("g2n_time_series") or {}).items()):
                ws4.cell(row=row_idx, column=1, value=sc.get("name"))
                ws4.cell(row=row_idx, column=2, value=cd.get("country_code"))
                ws4.cell(row=row_idx, column=3, value=int(year_str))
                ws4.cell(row=row_idx, column=4, value=g2n_val)
                row_idx += 1
    _autofit(ws4)

    # ── Sheet 5: Levers ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Levers")
    _header_row(ws5, LEVER_COLS)
    row_idx = 2
    for sc in scenario_configs:
        levers = sc.get("levers") or {}
        name = sc.get("name")
        for country in levers.get("withdrawals") or []:
            for col_idx, v in enumerate([name, "withdrawal", country, True], start=1):
                ws5.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
        for country, pct in (levers.get("price_floors") or {}).items():
            for col_idx, v in enumerate([name, "price_floor", country, pct], start=1):
                ws5.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
        for country, year in (levers.get("delayed_launches") or {}).items():
            for col_idx, v in enumerate([name, "delayed_launch", country, year], start=1):
                ws5.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
        for lever in ("de_opt_in", "gr_clawback_stress"):
            for col_idx, v in enumerate([name, lever, "", levers.get(lever, False)], start=1):
                ws5.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
    _autofit(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_asset_pack(
    asset: dict[str, Any],
    packs: list[dict[str, Any]],
    generated_by: str,
    asset_meta: dict[str, Any] | None = None,
    scenario_configs: list[dict[str, Any]] | None = None,
) -> bytes:
    """
    ZIP containing one sub-folder per simulation for an asset.

    packs: list of {"scenario_name", "sim", "audit_json"} dicts.
    If asset_meta + scenario_configs are provided, also adds asset_full.xlsx
    at the ZIP root — the importable round-trip file.
    Raises ExportTooLargeError if total exceeds MAX_EXPORT_BYTES.
    """
    import json

    asset_name = (asset.get("name") or "asset").replace(" ", "_")
    generated_at = datetime.now(timezone.utc).isoformat()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        prefix = f"{asset_name}/"

        readme = (
            f"PRICING STAR — Asset Audit Pack\n"
            f"{'=' * 40}\n\n"
            f"Asset:      {asset.get('name')}\n"
            f"Generated:  {generated_at}\n"
            f"Simulations: {len(packs)}\n\n"
            f"Each sub-folder contains audit.json + prices.xlsx for one simulation.\n"
            + (
                f"\nasset_full.xlsx — importable round-trip file (re-upload to update this asset)\n"
                if asset_meta is not None else ""
            )
        )
        zf.writestr(f"{prefix}README.txt", readme)

        # Importable round-trip file
        if asset_meta is not None and scenario_configs is not None:
            full_xlsx = _build_asset_full_xlsx(asset_meta, scenario_configs)
            zf.writestr(f"{prefix}asset_full.xlsx", full_xlsx)

        for pack in packs:
            scenario_name = pack["scenario_name"]
            sim = pack["sim"]
            audit_json = pack["audit_json"]
            sim_id = sim.get("simulation_id") or sim.get("id") or "unknown"
            folder = f"{prefix}{scenario_name.replace(' ', '_')}_{str(sim_id)[:8]}/"

            xlsx_bytes = build_simulation_xlsx(asset, scenario_name, sim)
            zf.writestr(f"{folder}audit.json", json.dumps(audit_json, indent=2))
            zf.writestr(f"{folder}prices.xlsx", xlsx_bytes)

    raw = buf.getvalue()
    if len(raw) > MAX_EXPORT_BYTES:
        raise ExportTooLargeError(f"Pack size {len(raw):,} bytes exceeds limit {MAX_EXPORT_BYTES:,}")
    return raw


def build_gdpr_export(
    user: dict[str, Any],
    tenant: dict[str, Any],
    assets: list[dict[str, Any]],
    scenarios: list[dict[str, Any]],
    simulations: list[dict[str, Any]],
    api_keys: list[dict[str, Any]],
    audit_logs: list[dict[str, Any]],
) -> bytes:
    """
    GDPR Article 20 data portability export.

    ZIP structure:
      my-data/
        README.txt
        profile.json
        tenant.json
        assets.json
        scenarios.json
        simulations.json
        api_keys.json      (no key_hash — secret not included)
        audit_log.json     (last 500 entries for this tenant)
    """
    import json

    generated_at = datetime.now(timezone.utc).isoformat()

    readme = (
        f"PRICING STAR — Your Data Export (GDPR Article 20)\n"
        f"{'=' * 50}\n\n"
        f"Generated:  {generated_at}\n"
        f"User:       {user.get('name')} ({user.get('email_masked')})\n"
        f"Tenant:     {tenant.get('name')}\n\n"
        f"Contents\n"
        f"--------\n"
        f"  profile.json     — your user profile\n"
        f"  tenant.json      — your organisation record\n"
        f"  assets.json      — all assets in your tenant\n"
        f"  scenarios.json   — all scenarios\n"
        f"  simulations.json — all simulation results\n"
        f"  api_keys.json    — API keys (secret hashes excluded)\n"
        f"  audit_log.json   — last 500 audit log entries\n\n"
        f"For questions about your data, contact support@pricingstar.io\n"
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        prefix = "my-data/"
        zf.writestr(f"{prefix}README.txt", readme)
        zf.writestr(f"{prefix}profile.json", json.dumps(user, indent=2, default=str))
        zf.writestr(f"{prefix}tenant.json", json.dumps(tenant, indent=2, default=str))
        zf.writestr(f"{prefix}assets.json", json.dumps(assets, indent=2, default=str))
        zf.writestr(f"{prefix}scenarios.json", json.dumps(scenarios, indent=2, default=str))
        zf.writestr(f"{prefix}simulations.json", json.dumps(simulations, indent=2, default=str))
        zf.writestr(f"{prefix}api_keys.json", json.dumps(api_keys, indent=2, default=str))
        zf.writestr(f"{prefix}audit_log.json", json.dumps(audit_logs, indent=2, default=str))

    raw = buf.getvalue()
    if len(raw) > MAX_EXPORT_BYTES:
        raise ExportTooLargeError(f"Export size {len(raw):,} bytes exceeds limit {MAX_EXPORT_BYTES:,}")
    return raw
