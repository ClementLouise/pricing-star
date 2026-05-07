"""Unit tests for app.services.asset_import — parse, validate, template."""
from __future__ import annotations

import io
import uuid

import openpyxl
import pytest

from app.services.asset_import import (
    ImportMode,
    build_import_template,
    parse_xlsx,
)
from app.services.audit_pack import (
    COUNTRY_DATA_COLS,
    G2N_COLS,
    LEVER_COLS,
    OVERVIEW_KEYS,
    SCENARIO_COLS,
    _build_asset_full_xlsx,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_wb(sheets: dict[str, list[list]]) -> bytes:
    """Build an in-memory XLSX from {sheet_name: rows}."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _overview_rows(overrides: dict | None = None) -> list[list]:
    defaults: dict = {
        "name": "Test Asset",
        "therapeutic_area": "Oncology",
        "modality": "Small Molecule",
        "indication": "NSCLC",
        "us_list_price": 250000,
        "us_net_share": 0.70,
        "launch_year": 2027,
        "peak_year": 2031,
        "loe_year": 2040,
        "cogs_percent": 0.15,
        "discount_rate": 0.10,
        "us_patient_population": 40000,
        "ex_us_patient_population": 25000,
        "peak_capture_rate": 0.50,
        "part_b_share": 0.50,
        "ramp_years": 4,
    }
    if overrides:
        defaults.update(overrides)
    rows: list[list] = [["key", "value"]]
    for k, v in defaults.items():
        rows.append([k, v])
    return rows


def _scenario_row(name: str = "S1", is_baseline: bool = True) -> list:
    return [
        name, "desc", is_baseline,
        False, None, 0.07,
        False, None, False, None,
        False, None, False, None,
        True, 10,
    ]


def _country_row(scenario_name: str = "S1", country: str = "DE") -> list:
    return [scenario_name, country, 150000, 120000, 0.4, True, 2027, False, 0.80]


def _minimal_xlsx(scenario_name: str = "S1") -> bytes:
    return _make_wb({
        "Overview": _overview_rows(),
        "Scenarios": [SCENARIO_COLS[:], _scenario_row(scenario_name)],
        "CountryData": [COUNTRY_DATA_COLS[:], _country_row(scenario_name)],
        "G2N_TimeSeries": [G2N_COLS[:]],   # header only — empty
        "Levers": [LEVER_COLS[:]],          # header only — empty
    })


# ── CREATE_NEW ────────────────────────────────────────────────────────────────

class TestParseCreateNew:
    def test_valid_minimal_file_no_errors(self):
        plan = parse_xlsx(_minimal_xlsx(), ImportMode.CREATE_NEW)
        assert not plan.errors
        assert plan.parsed is not None
        assert plan.summary["scenario_count"] == 1

    def test_parses_asset_name(self):
        plan = parse_xlsx(_minimal_xlsx(), ImportMode.CREATE_NEW)
        assert plan.parsed.asset_payload.name == "Test Asset"

    def test_invalid_xlsx_bytes(self):
        plan = parse_xlsx(b"not an xlsx file", ImportMode.CREATE_NEW)
        assert any(e.code == "INVALID_XLSX" for e in plan.errors)

    def test_missing_overview_sheet_raises_missing_sheet(self):
        xlsx = _make_wb({
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "MISSING_REQUIRED_SHEET" for e in plan.errors)

    def test_empty_name_raises_missing_required_field(self):
        xlsx = _make_wb({
            "Overview": _overview_rows({"name": ""}),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "MISSING_REQUIRED_FIELD" for e in plan.errors)

    def test_invalid_country_code(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row(country="XX")],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "INVALID_COUNTRY_CODE" for e in plan.errors)

    def test_multiple_baselines(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [
                SCENARIO_COLS[:],
                _scenario_row("S1", is_baseline=True),
                _scenario_row("S2", is_baseline=True),
            ],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row("S1")],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "MULTIPLE_BASELINES" for e in plan.errors)

    def test_unknown_scenario_ref_in_country_data(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row("S1")],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row("NONEXISTENT")],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "UNKNOWN_SCENARIO_REF" for e in plan.errors)

    def test_invalid_g2n_value_above_one(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
            "G2N_TimeSeries": [G2N_COLS[:], ["S1", "DE", 2027, 1.5]],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "INVALID_G2N" for e in plan.errors)

    def test_g2n_year_out_of_range(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
            "G2N_TimeSeries": [G2N_COLS[:], ["S1", "DE", 2000, 0.85]],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "G2N_YEAR_OUT_OF_RANGE" for e in plan.errors)

    def test_invalid_lever_type(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
            "Levers": [LEVER_COLS[:], ["S1", "bad_lever_type", "DE", 1]],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert any(e.code == "INVALID_LEVER_TYPE" for e in plan.errors)

    def test_real_pricing_data_emits_warning_not_error(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
            # list_price already 150k < 100k threshold — use a larger one
        })
        big_row = ["S1", "DE", 500_001, 400000, 0.3, True, 2027, False, 0.80]
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], big_row],
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert not plan.errors
        assert any(w.code == "REAL_PRICING_DETECTED" for w in plan.warnings)

    def test_no_scenarios_emits_warning_not_error(self):
        xlsx = _make_wb({
            "Overview": _overview_rows(),
            "Scenarios": [SCENARIO_COLS[:]],       # header only
            "CountryData": [COUNTRY_DATA_COLS[:]],  # header only
        })
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert not plan.errors
        assert any(w.code == "NO_SCENARIOS" for w in plan.warnings)


# ── UPDATE_EXISTING ───────────────────────────────────────────────────────────

class TestParseUpdateExisting:
    def test_missing_asset_id_raises_error(self):
        xlsx = _make_wb({
            "Overview": [["key", "value"], ["name", "Test Asset"]],
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
        })
        plan = parse_xlsx(xlsx, ImportMode.UPDATE_EXISTING)
        assert any(e.code == "MISSING_ASSET_ID" for e in plan.errors)

    def test_asset_id_mismatch_raises_error(self):
        asset_id = uuid.uuid4()
        wrong_id = uuid.uuid4()
        rows = _overview_rows()
        rows.append(["asset_id", str(asset_id)])
        xlsx = _make_wb({
            "Overview": rows,
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
        })
        plan = parse_xlsx(xlsx, ImportMode.UPDATE_EXISTING, target_asset_id=wrong_id)
        assert any(e.code == "ASSET_ID_MISMATCH" for e in plan.errors)

    def test_valid_with_matching_asset_id(self):
        asset_id = uuid.uuid4()
        rows = _overview_rows()
        rows.append(["asset_id", str(asset_id)])
        xlsx = _make_wb({
            "Overview": rows,
            "Scenarios": [SCENARIO_COLS[:], _scenario_row()],
            "CountryData": [COUNTRY_DATA_COLS[:], _country_row()],
        })
        plan = parse_xlsx(xlsx, ImportMode.UPDATE_EXISTING, target_asset_id=asset_id)
        assert not plan.errors
        assert plan.parsed is not None
        assert plan.parsed.asset_id == asset_id


# ── Round-trip ────────────────────────────────────────────────────────────────

class TestRoundTrip:
    def test_exported_xlsx_parses_without_errors(self):
        """asset_full.xlsx built by audit_pack must re-parse cleanly."""
        from datetime import datetime, timezone

        asset_meta = {
            "asset_id": str(uuid.uuid4()),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "name": "Round Trip Drug",
            "therapeutic_area": "Oncology",
            "modality": "mAb",
            "indication": "NSCLC",
            "us_list_price": 180000,
            "us_net_share": 0.72,
            "launch_year": 2026,
            "peak_year": 2030,
            "loe_year": 2038,
            "cogs_percent": 0.14,
            "discount_rate": 0.09,
            "us_patient_population": 35000,
            "ex_us_patient_population": 20000,
            "peak_capture_rate": 0.45,
            "part_b_share": 0.55,
            "ramp_years": 3,
        }
        scenario_configs = [{
            "name": "Base Case",
            "description": "baseline",
            "is_baseline": True,
            "regulations": {
                "generous": {"active": False, "year": None, "medicaid_share": 0.07},
                "guard": {"active": False, "year": None, "submit_method_ii": False, "phase_in": None},
                "globe": {"active": False, "year": None, "submit_method_ii": False, "phase_in": None},
            },
            "cascade_config": {"enabled": True, "max_iter": 10},
            "levers": {
                "withdrawals": [], "price_floors": {}, "delayed_launches": {},
                "de_opt_in": False, "gr_clawback_stress": False,
            },
            "country_data": [{
                "country_code": "DE",
                "list_price": 90000,
                "net_price": 75000,
                "volume": 0.3,
                "launched": True,
                "launch_year": 2026,
                "withdrawn": False,
                "g2n_ratio": 0.83,
                "g2n_time_series": {"2027": 0.81, "2028": 0.79},
            }],
        }]

        xlsx_bytes = _build_asset_full_xlsx(asset_meta, scenario_configs)
        plan = parse_xlsx(xlsx_bytes, ImportMode.CREATE_NEW)

        assert not plan.errors, f"Round-trip errors: {[e.code for e in plan.errors]}"
        assert plan.parsed is not None
        assert plan.parsed.asset_payload.name == "Round Trip Drug"
        assert len(plan.parsed.scenarios) == 1
        assert plan.parsed.scenarios[0].name == "Base Case"
        assert len(plan.parsed.scenarios[0].country_data) == 1


# ── Template generation ───────────────────────────────────────────────────────

class TestBuildImportTemplate:
    def test_returns_valid_xlsx_with_five_sheets(self):
        xlsx = build_import_template("Test Drug")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx))
        for sheet in ("Overview", "Scenarios", "CountryData", "G2N_TimeSeries", "Levers"):
            assert sheet in wb.sheetnames

    def test_template_parses_without_errors(self):
        xlsx = build_import_template("My Test Asset")
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert not plan.errors
        assert plan.parsed is not None
        assert plan.parsed.asset_payload.name == "My Test Asset"

    def test_template_has_sample_scenario(self):
        xlsx = build_import_template()
        plan = parse_xlsx(xlsx, ImportMode.CREATE_NEW)
        assert not plan.errors
        assert plan.summary["scenario_count"] == 1
